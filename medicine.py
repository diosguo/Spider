from bs4 import BeautifulSoup
import requests
import xlwt
from openpyxl import *

def outToExcel(datas):
    wb=Workbook()
    ws=wb.create_sheet('sheet1')
    ws.cell(row=1,column=1,value='商品名称')
    ws.cell(row=1,column=2,value='通用名称')
    ws.cell(row=1,column=3,value='生产企业')
    ws.cell(row=1,column=4, value='批准文号')
    ws.cell(row=1, column=5, value='适应症')
    ws.cell(row=1, column=6, value='剂型')
    ws.cell(row=1, column=7, value='规格')
    ws.cell(row=1, column=8, value='用法用量')
    ws.cell(row=1, column=9, value='不良反应')
    ws.cell(row=1, column=10, value='有效期')
    propMap={
        'name':0,
        'gname':1,
        'org':2,
        'license':3,
        'application':4,
        'type':5,
        'spec':6,
        'method':7,
        'effect':8,
        'period':9
    }
    row=2
    for data in datas:
        keys=data.keys()
        for key in keys:
            ws.cell(row=row,column=propMap[key]+1,value=data[key])
        row+=1

    wb.save('meddddd.xlsx')

def getInfo(url):
    data={}
    wb_data = requests.get(url)
    soup = BeautifulSoup(wb_data.text, 'lxml')
    medicine = soup.select('#wrap990list1 > ul > li')
    string=str(medicine)
    if(string.find('不良反应')>0):
        data['effect']=string[string.find('不良反应'):string.find('<',string.find('不良反应'))]
    if (string.find('商品名称') > 0):
        data['name'] = string[string.find('商品名称')+5:string.find('<', string.find('商品名称'))]
    if(string.find('规格')>0):
        data['spec']=string[string.find('规格')+3:string.find('<',string.find('规格'))]
    if (string.find('生产企业') > 0):
        data['org'] = string[string.find('生产企业')+5:string.find('<', string.find('生产企业'))]
    if(string.find('通用名称')>0):
        data['gname']=string[string.find('通用名称')+5:string.find('<',string.find('通用名称'))]
    if (string.find('有效期') > 0):
        data['period'] = string[string.find('有效期')+4:string.find('<', string.find('有效期'))]
    if(string.find('剂型')>0):
        data['type']=string[string.find('剂型')+3:string.find('<',string.find('剂型'))]
    if (string.find('批准文号') > 0):
        data['license'] = string[string.find('批准文号')+5:string.find('<', string.find('批准文号'))]
    if (string.find('适应症') > 0):
        data['application'] = string[string.find('适应症')+9:string.find('<', string.find('适应症'))]
    if (string.find('用法用量') > 0):
        data['method'] = string[string.find('用法用量')+5:string.find('<', string.find('用法用量'))]
    return data

def run(url):
    print('in run:'+str(url))
    wb_data=requests.get(url)
    soup=BeautifulSoup(wb_data.text,'lxml')
    medicines=soup.select('ul.Productlist p.pic > a')
    datas=[]
    for med in medicines:
        href=med.get('href')
        datas.append(getInfo(href[:href.find('?')]))
    return datas

def page(url):
    datas=[]
    datas.extend(run(url))
    while(True):
        print('in page:'+str(url))
        wb_data=requests.get(url)
        soup=BeautifulSoup(wb_data.text,'lxml')
        nextpage = soup.select('a.Ynext')
        print('next page;'+str(nextpage))
        if str(nextpage).find('disable') > 0:
            return datas
        else:
            try:
                url='http://search.360kad.com'+nextpage[0].get('href')
            except IndexError:
                return datas
            datas.extend(run(url))
    return datas
def pa():
    url = 'http://www.360kad.com'
    wb_data=requests.get(url)
    soup = BeautifulSoup(wb_data.text, 'lxml')
    index=soup.select('#lNav_lists > li:nth-of-type(1) > div.lNav_pop > dl > dt > a')
    lenofindex=len(index)
    datas=[]
    for j in range(lenofindex):
        href = str(index[j].get('href'))
        print('in pa:'+str(href))
        datas.extend(page(href))
        print(len(datas))
        print(datas)
    return datas

if __name__ =='__main__':
    datas=pa()
    outToExcel(datas)